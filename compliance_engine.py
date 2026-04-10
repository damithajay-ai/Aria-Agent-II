"""
99x Compliance Engine
All calculation logic lives here on the backend.
Accepts file bytes directly (no disk I/O needed).
"""

from openpyxl import load_workbook
from collections import defaultdict
from io import BytesIO
import re


def _wb(file_bytes):
    return load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

def _rows(wb, sheet_name=None):
    ws = wb[sheet_name] if sheet_name else wb.active
    return list(ws.iter_rows(values_only=True))

def _find_hdr(rows, keyword):
    for i, row in enumerate(rows):
        if row and any(c and keyword.lower() in str(c).lower() for c in row):
            return i
    return -1

def _col(h, name):
    for i, c in enumerate(h):
        if c and name.lower() in str(c).lower():
            return i
    return -1

def norm(s):
    if not s: return ""
    return str(s).replace("\t", " ").replace("\s+", " ").strip().lower()


def parse_timesheet(file_bytes: bytes):
    wb   = _wb(file_bytes)
    rows = _rows(wb)
    hi   = _find_hdr(rows, "userid")
    if hi < 0:
        raise ValueError("Could not find header row in timesheet file.")

    h      = [str(c).lower().strip() if c else "" for c in rows[hi]]
    i_id   = h.index("userid")
    i_name = next(i for i, c in enumerate(h) if c == "name")
    i_bill = next(i for i, c in enumerate(h) if "billable" in c)
    i_tot  = next(i for i, c in enumerate(h) if c == "total workhours")
    i_acct = next((i for i, c in enumerate(h) if c == "account"), -1)

    ts_map  = {}
    name_map = {}
    for row in rows[hi + 1:]:
        if not row or not row[i_id]: continue
        uid  = str(row[i_id]).strip()
        nm   = norm(row[i_name])
        bill = float(row[i_bill] or 0)
        tot  = float(row[i_tot]  or 0)
        acct = str(row[i_acct]).strip() if i_acct >= 0 and row[i_acct] else ""

        if uid not in ts_map:
            ts_map[uid] = {
                "name":     str(row[i_name] or "").replace("\t", " ").strip(),
                "billable": 0.0,
                "total":    0.0,
                "accounts": [],
            }
        # Same total hours repeated per project row — take MAX not SUM
        ts_map[uid]["total"]   = max(ts_map[uid]["total"],   tot)
        ts_map[uid]["billable"] = max(ts_map[uid]["billable"], bill)
        if acct and acct not in ts_map[uid]["accounts"]:
            ts_map[uid]["accounts"].append(acct)
        if nm and nm not in name_map:
            name_map[nm] = uid

    return ts_map, name_map


def parse_staff(file_bytes: bytes):
    wb   = _wb(file_bytes)
    rows = _rows(wb)
    hi   = _find_hdr(rows, "employee")
    if hi < 0:
        raise ValueError("Could not find header row in staff profile file.")

    h      = [str(c).lower().strip() if c else "" for c in rows[hi]]
    i_num  = next(i for i, c in enumerate(h) if "employee" in c and "number" in c)
    i_name = next(i for i, c in enumerate(h) if c == "name")

    staff = {}
    for row in rows[hi + 1:]:
        if not row or row[i_num] is None: continue
        num = str(row[i_num]).strip()
        if not num: continue
        staff[num] = str(row[i_name] or "").replace("\s+", " ").strip()
    return staff


def parse_leave(file_bytes: bytes, month: str):
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet_name = next(
        (s for s in wb.sheetnames if "edit" in s.lower()), wb.sheetnames[0]
    )
    rows = _rows(wb, sheet_name)
    hi   = _find_hdr(rows, "number")
    if hi < 0:
        raise ValueError("Could not find header row in leave file.")

    h        = [str(c).lower().strip() if c else "" for c in rows[hi]]
    i_num    = next(i for i, c in enumerate(h) if c == "number")
    month_abbr = month[:3].lower()
    i_month  = next((i for i, c in enumerate(h) if c == month_abbr), -1)

    leave_map = defaultdict(float)
    last_emp  = None
    for row in rows[hi + 1:]:
        if not row or all(c is None for c in row): continue
        emp = str(row[i_num]).strip() if row[i_num] else None
        if emp and not emp.startswith("="):
            last_emp = emp
        elif not emp or emp.startswith("="):
            emp = last_emp
        if not emp: continue
        days = float(row[i_month] or 0) if i_month >= 0 and row[i_month] else 0
        leave_map[emp] += days

    return dict(leave_map)


def parse_holidays(file_bytes: bytes, month: str, region: str = "SL"):
    wb   = _wb(file_bytes)
    rows = _rows(wb)
    hi   = _find_hdr(rows, "month")
    if hi < 0:
        raise ValueError("Could not find header row in holiday calendar file.")

    h        = [str(c).lower().strip() if c else "" for c in rows[hi]]
    i_month  = next(i for i, c in enumerate(h) if c == "month")
    i_wd     = next(i for i, c in enumerate(h) if "working" in c)
    i_hol    = next(i for i, c in enumerate(h) if "holiday" in c)
    i_region = next((i for i, c in enumerate(h) if "region" in c), -1)

    for row in rows[hi + 1:]:
        if not row: continue
        rm = str(row[i_month] or "").strip().lower()
        rr = str(row[i_region] or "").strip() if i_region >= 0 else region
        if rm == month.lower() and rr == region:
            wd  = float(row[i_wd]  or 21)
            hol = float(str(row[i_hol] or 1).replace("*", "") or 1)
            return wd, hol

    return 21.0, 1.0


def build_report(ts_map, name_map, staff_map, leave_map, working_days, holidays, month):
    records = []
    for emp_num, emp_name in staff_map.items():
        nm  = norm(emp_name)
        uid = name_map.get(nm)
        if not uid:
            uid = next(
                (name_map[k] for k in name_map if k in nm or nm in k), None
            )
        ts = ts_map.get(uid) if uid else None

        # Match leave by emp_num — handle numeric/string mismatch
        leave_applied = leave_map.get(emp_num, 0)
        if not leave_applied:
            try:
                leave_applied = leave_map.get(str(int(float(emp_num))), 0)
            except (ValueError, TypeError):
                pass

        total_leave_days  = round((leave_applied + holidays) * 10) / 10
        leave_hours       = round(total_leave_days * 8 * 10) / 10
        required_hours    = round(working_days * 8 * 10) / 10
        total_hours       = round(ts["total"]   if ts else 0, 2)
        billable_hours    = round(ts["billable"] if ts else 0, 2)
        contractual_hours = round((total_hours + leave_hours) * 10) / 10
        compliance        = round((contractual_hours / required_hours) * 100) if required_hours > 0 else 0
        accounts          = ts["accounts"] if ts and ts["accounts"] else []

        records.append({
            "empId":            emp_num,
            "name":             emp_name,
            "accounts":         accounts,
            "billableHours":    billable_hours,
            "totalHours":       total_hours,
            "totalLeaveApplied":leave_applied,
            "totalHolidays":    holidays,
            "totalLeaveDays":   total_leave_days,
            "leaveHours":       leave_hours,
            "contractualHours": contractual_hours,
            "requiredHours":    required_hours,
            "compliance":       compliance,
        })

    records.sort(key=lambda r: r["name"])
    return records
